import os
from openpyxl import load_workbook
from PyPDF2 import PdfReader, PdfWriter
from datetime import datetime
from tkinter import filedialog, Tk, messagebox

def split_pdf_by_pairs(pdf_path, excel_path, output_folder):
    if not pdf_path or not excel_path or not output_folder:
        return "Отменено пользователем"
    
    log_file_path = f'log_{datetime.now().strftime("%Y_%m_%d_%H%M%S")}.txt'
    with open(log_file_path, 'w') as log_file:
        if not os.path.exists(pdf_path):
            log_file.write(f"Файл PDF {pdf_path} не существует.\n")
            return "Неудача: Файл PDF не существует."
        
        if not os.path.exists(excel_path):
            log_file.write(f"Файл Excel {excel_path} не существует.\n")
            return "Неудача: Файл Excel не существует."
        
        wb = load_workbook(filename=excel_path, read_only=True)
        ws = wb.active

        names = []
        pages = []

        for row in ws.iter_rows(min_row=2):
            name_parts = [row[1].value, row[2].value, row[3].value]
            name = f"{name_parts[0]}_{name_parts[1]}_{name_parts[2]}"
            page_count = row[4].value
            
            # Check if all parts of the name are not None
            if None in name_parts or page_count is None or not isinstance(page_count, int) or page_count <= 0:
                if not all(part is None for part in name_parts):
                    log_file.write(f"Неверное значение страниц: {page_count} для {name}. Пропуск.\n")
                continue

            names.append(name)
            pages.append(page_count)

        if not names or not pages:
            log_file.write("Нет допустимых данных для обработки.\n")
            return "Неудача: Нет допустимых данных для обработки."

        with open(pdf_path, 'rb') as f:
            reader = PdfReader(f)
            num_pages = len(reader.pages)

            if num_pages != sum(pages):
                difference = abs(num_pages - sum(pages))
                log_file.write(f"Общее количество страниц в PDF ({num_pages}) не соответствует общему количеству страниц, указанному в файле Excel ({sum(pages)}). Разница составляет {difference} страниц.\n")
                return f"Неудача: Общее количество страниц в PDF не соответствует общему количеству страниц, указанному в файле Excel. Разница составляет {difference} страниц."
            
            os.makedirs(output_folder, exist_ok=True)
            file_counter = 0
            page_counter = 0

            while page_counter < num_pages and file_counter < len(names):
                writer = PdfWriter()
                name = names[file_counter]
                num_pages_to_save = pages[file_counter]

                for _ in range(num_pages_to_save):
                    if page_counter < num_pages:
                        writer.add_page(reader.pages[page_counter])
                        page_counter += 1

                output_pdf_path = os.path.join(output_folder, f'{name}.pdf')
                with open(output_pdf_path, 'wb') as output_pdf:
                    writer.write(output_pdf)

                log_file.write(f"Создан PDF-файл: {output_pdf_path}\n")
                file_counter += 1

        return f"Успех: Программа успешно завершила работу. Создано {file_counter} файлов."

def ask_file_path(title, filetypes):
    while True:
        file_path = filedialog.askopenfilename(title=title, filetypes=filetypes)
        if not file_path:
            if messagebox.askyesno("PDF Convert", "Выбор файла был отменен. Вы хотите попробовать снова?"):
                continue
            else:
                return None
        return file_path

def main():
    root = Tk()
    root.withdraw()
    messagebox.showinfo("PDF Convert", "Здесь могла быть ваша реклама)) © Белоусов А.В.")
    
    pdf_file_path = ask_file_path("Выберите PDF файл", [("PDF files", "*.pdf")])
    if not pdf_file_path:
        messagebox.showinfo("PDF Convert", "Выбор PDF файла был отменен")
        root.destroy()
        return
    
    excel_file_path = ask_file_path("Выберите Excel файл", [("Excel files", "*.xlsx")])
    if not excel_file_path:
        messagebox.showinfo("PDF Convert", "Выбор Excel файла был отменен")
        root.destroy()
        return
    
    output_folder_path = filedialog.askdirectory(title="Выберите папку для сохранения PDF файлов")
    if not output_folder_path:
        messagebox.showinfo("PDF Convert", "Выбор папки для сохранения PDF файлов был отменен")
        root.destroy()
        return
    
    result = split_pdf_by_pairs(pdf_file_path, excel_file_path, output_folder_path)
    messagebox.showinfo("Результат © Белоусов А.В.", result)
    root.destroy()

if __name__ == "__main__":
    main()
